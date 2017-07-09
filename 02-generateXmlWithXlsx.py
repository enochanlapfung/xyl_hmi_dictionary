#!/usr/bin/python

#-------------------------------------------------------------------------------
# XML dictionary generator
# Uses a master dictionary (xlsx) and a .def file containing the required
# strings in order to produce XML dictionaries in several languages
#-------------------------------------------------------------------------------

import xml.etree.ElementTree as ET
import xml.dom.minidom
import os
import parse
import re
from openpyxl import Workbook
from openpyxl import load_workbook

# Title of sheet 1 in the xlsx work book
# Ideally it should be "english"
SHEET1_TITLE = "MasterDictionary"

# Language code to language mapping
# This dictionary's values will be used to generate XML file names
# e.g "english.xml"
languages = {
    SHEET1_TITLE:"english",
    "hr-HR":"croatian",
    "zh-CN":"chinese",
    "da-DK":"danish",
    "nl-NL":"dutch",
    "fi-FI":"finnish",
    "fr-FR":"french",
    "de-DE":"german",
    "hu-HU":"hungarian",
    "it-IT":"italian",
    "no-NO":"norwegian",
    "pl-PL":"polish",
    "pt-PT":"portuguese",
    "es-ES":"spanish",
    "sv-SE":"swedish",
    "tr-TR":"turkish",
}

# Dictionary which will contain a dictionary for each language
# Each of those dictionaries will contain UID to string pairs
dictionary = {}

outputDirectory = "dictionaries"

#-------------------------------------------------------------------------------
# Parsing def file
#-------------------------------------------------------------------------------
uids = []
deffile = open("example.def", 'r')
for line in deffile:
    # Extracting ID and uID
    result = parse.parse("STRING_DEF({}, {:d})", line)
    idString    = result.fixed[0]
    uid         = result.fixed[1]
    uids.append(uid)







# Loading master dictionary workbook
wb = load_workbook("MasterDictionary.xlsx")

# Iterating through each of the sheets
for sheet in wb:

    print "Parsing sheet: " + sheet.title

    currentDict = dictionary[languages[sheet.title]] = {}

    # Obtain only the rows that have a uid in uids
    for uid in uids:

        # Obtain uid and string
        row = uid + 1
        sheetUid    = sheet.cell(row=row, column=1).value
        sheetString = sheet.cell(row=row, column=2).value
        if not (re.search("^u_", sheetUid) and len(sheetUid) == 8):
            # SheetUid is defined on the first sheet, same cell
            sheetUid = wb[SHEET1_TITLE].cell(row=row, column=1).value

        # Obtain integer uid
        uidInt = parse.parse("u_{:d}", sheetUid).fixed[0]

        # Check whether integer uid is valid
        if uidInt not in uids:
            print "Invalid UID found"
            exit()

        # Convert to UTF-8
        sheetUidUtf8    = unicode(sheetUid).encode('utf-8')
        sheetStringUtf8 = unicode(sheetString).encode('utf-8')

        # print sheetUidUtf8, sheetStringUtf8

        # Add to dictionary
        currentDict[uidInt] = sheetString

    # keys = currentDict.keys()
    # keys.sort()
    # print keys
    # print ""






for language in dictionary:

    print "Generating XML for ", language

    # Building tree from uids and master dictionary
    # Root element
    root = ET.Element("concept", {"id":"c_dictionary", "xml:lang":"en-US"})
    # Title element
    title = ET.SubElement(root, "title")
    title.text = "CompactController dictionary"
    # Conbody element
    conbody = ET.SubElement(root, "conbody")

    # Creating sections

    uidKeys = dictionary[language].keys()
    uidKeys.sort()
    for uidInt in uidKeys:
        section = ET.SubElement(conbody, "section")
        menucascade = ET.SubElement(section, "menucascade")

        # Master dictionary gives uid with an underscore (_), but we required
        # UIDs to have dashes (-)
        uid = "u-%06d" % (uidInt)
        uicontrol = ET.SubElement(menucascade, "uicontrol", {"id":uid})
        uicontrol.text = dictionary[language][uidInt]

    # Create the tree
    # tree = ET.ElementTree(root)


    # Create pretty XML output


    roughString = ET.tostring(root, encoding="UTF-8")
    reparse = xml.dom.minidom.parseString(roughString)
    doc = reparse.toprettyxml(indent="\t")
    # Write the pretty XML to a file
    if not os.path.exists(outputDirectory):
        os.mkdir(outputDirectory)


    docUtf8 = unicode(doc).encode('utf8')

    fileOutput = open(outputDirectory + "\\" + language + ".xml", 'w')
    fileOutput.write(docUtf8)
    fileOutput.close()
